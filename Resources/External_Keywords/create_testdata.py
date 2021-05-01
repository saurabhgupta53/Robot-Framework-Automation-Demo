from faker import Faker
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet("LoginData")

fkr = Faker()
for i in range(1, 11):
    ws.cell(row=i, column=1).value = fkr.first_name()
    ws.cell(row=i, column=2).value = fkr.password()


wb.save("C:/Users/euagsup/PycharmProjects/Learnings/RobotFW/TestData/TestData1.xlsx")

