import openpyxl

wb = openpyxl.load_workbook("C:/Users/euagsup/PycharmProjects/Learnings/RobotFW/TestData/TestData1.xlsx")
# ws = wb["LoginData"]


def count_rows_in_excel(sheet_name):
    ws = wb[sheet_name]
    return ws.max_row


def fetch_data_from_excel(sheet_name, row_ct, cell):
    ws = wb[sheet_name]
    cell_data = ws.cell(int(row_ct), int(cell))
    return cell_data.value


# print(fetch_data_from_excel("LoginData", 4, 2))
